from __future__ import unicode_literals
import frappe
import os
import frappe.utils
import requests
import json
from frappe.utils import cint, cstr, flt, getdate, nowdate
from frappe.desk import query_report
from six import string_types
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
from openpyxl.styles.borders import Border,Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from jinja2 import Environment, BaseLoader, PackageLoader,FileSystemLoader
# from email.MIMEBase import MIMEBase
import base64
import logging

# logging.basicConfig(filename='outstanding_mail_log.log', level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S') # file mode --> filemode='w'

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
# print('location: ',__location__) #location:  /home/dbtpl/frappe-bench/apps/remainders/remainders/remainder_automation

@frappe.whitelist()
def generate_customer_outstanding_data(customer_name):
	print('\n\----------> Customer Name: ', customer_name)
	customer_filter,last_row = get_data(customer_name)

	thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

	if len(customer_filter) != 0 :
		wb = load_workbook(os.path.join(__location__, 'Outstanding_remainder_template.xlsx'))
		template_sheet = wb['Remainder']
		template_sheet['A14'] = customer_name

		for row in range(18,len(customer_filter)+18):
			i = 1
			for key in customer_filter[row-18].keys():

				# This code for applying border to cell
				for i1 in range(row,row+1):
					for j1 in range(1,11):
						ch = chr(64+j1)
						template_sheet[ch+(str(i1))].border = thin_border

				if key == 'name':
					_ = template_sheet.cell(column=1, row=row, value= customer_filter[row-18][key])
				elif key == 'posting_date':
					_ = template_sheet.cell(column=2, row=row, value= customer_filter[row-18][key])
				elif key == 'po_no':
					_ = template_sheet.cell(column=3, row=row, value= customer_filter[row-18][key])
				elif key == 'po_date':
					_ = template_sheet.cell(column=4, row=row, value= customer_filter[row-18][key])
				elif key == 'due_date':
					_ = template_sheet.cell(column=5, row=row, value= customer_filter[row-18][key])
				elif key == 'age':
					_ = template_sheet.cell(column=6, row=row, value= customer_filter[row-18][key])
				elif key == 'base_rounded_total':
					_ = template_sheet.cell(column=7, row=row, value= customer_filter[row-18][key])
				elif key == 'paid_amount':
					_ = template_sheet.cell(column=8, row=row, value= customer_filter[row-18][key])
				elif key == 'cn_amount':
					_ = template_sheet.cell(column=9, row=row, value= customer_filter[row-18][key])
				elif key == 'outstanding_amount':
					_ = template_sheet.cell(column=10, row=row, value= customer_filter[row-18][key])
					_.font = Font(bold=True,size=10)

		#Blank Line
		cell1 = f"A{str(row+1)}:F{str(row+1)}"
		template_sheet.merge_cells(f'{cell1}')
		template_sheet.row_dimensions[row+1].height = 5

		template_sheet['A'+(str(row+2))] = "Grand Total"
		template_sheet['A'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['A'+(str(row+2))].alignment = Alignment(horizontal='center')
		template_sheet['A'+(str(row+2))].border = thin_border
		cell1 = f"B{str(row+2)}:F{str(row+2)}"
		template_sheet.merge_cells(f'{cell1}')

		# Merge Cell Border
		template_sheet['B'+(str(row+2))].border = thin_border
		template_sheet['C'+(str(row+2))].border = thin_border
		template_sheet['D'+(str(row+2))].border = thin_border
		template_sheet['E'+(str(row+2))].border = thin_border
		template_sheet['F'+(str(row+2))].border = thin_border
		
		template_sheet['G'+(str(row+2))] = last_row[9]
		template_sheet['G'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['G'+(str(row+2))].border = thin_border
		sum_inv_amt= last_row[9]

		template_sheet['H'+(str(row+2))] = last_row[10]
		template_sheet['H'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['H'+(str(row+2))].border = thin_border
		sum_paid_amt= last_row[10]

		template_sheet['I'+(str(row+2))] = last_row[11]
		template_sheet['I'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['I'+(str(row+2))].border = thin_border
		sum_cn_amt= last_row[11]

		template_sheet['J'+(str(row+2))] = last_row[12]
		template_sheet['J'+(str(row+2))].font = Font(bold=True,size=10)
		template_sheet['J'+(str(row+2))].border = thin_border
		sum_outstanding_amt= last_row[12]

		template_sheet.row_dimensions[row+2].height = 15

		wb.save(os.path.join(__location__, 'outstanding_remainder.xlsx'))
		wb.close()
		# print('\n\nreturn: ',customer_filter,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt)

		return customer_filter,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt

def get_customer():
    data = frappe.db.sql(f"""
                            SELECT customer_name,billing_email_id as email_id
                            FROM `tabCustomer`
			    			WHERE disabled = 0
                            """,as_dict=1)
    return data

def get_data(cust):
	outstanding = query_report.run(report_name = "Accounts Receivable",filters= {"company":"Dhupar Brothers Trading Pvt. Ltd.","party": (cust,),"report_date":datetime.today().strftime('%Y-%m-%d'),"ageing_based_on":"Posting Date","range1":30,"range2":60,"range3":90,"range4":120},ignore_prepared_report= True)
	outstanding = outstanding['result']

	final_data = []
	last_row = None
	for i in outstanding:
		if type(i) != list:
			if 'po_no' in i:
				po_date = frappe.db.sql(f"""
											SELECT name,po_date
											FROM `tabSales Invoice`
											WHERE name = '{i.voucher_no}'
									""",as_dict=1)[0]
				i['po_date'] = po_date.po_date

			final_data.append({
				"name": i.voucher_no,
				"posting_date":i.posting_date,
				"po_no":i.po_no,
				"po_date":i.po_date,
				"due_date":i.due_date,
				"age":i.age,
				"base_rounded_total": i.invoice_grand_total,
				"paid_amount": i.paid,
				"cn_amount": i.credit_note,
				"outstanding_amount": i.outstanding
				})
		else:
			if i[12] > 0:
				last_row = i
			else:
				final_data = []
	# print('final_data,last_row: ',final_data,last_row)
	return final_data,last_row


# CALL --> schedular_event(weekly) or TESTING --> by using bench console
@frappe.whitelist()
def send_outstanding_mail(customer,email):
	# customer = get_customer()

	cust = customer
	email = email.split(',') # If multiple email then we split here (E.g --> ['aonpowerline@rediffmail.com', 'mahalaxmielectricals9922@gmail.com'])

	customer = []
	for i in email:
		data = {'customer_name': cust,'email_id':i}
		customer.append(data)


	if len(customer) != 0:
		for i in customer:
			ret_data = generate_customer_outstanding_data(i['customer_name'])
			if ret_data != None :
				customer_outstanding_data,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt = ret_data
				# print('Customer return data: ',customer_outstanding_data,sum_inv_amt,sum_paid_amt,sum_cn_amt,sum_outstanding_amt)

				attachments = [{
						'fname': "outstanding_remainder.xlsx",
						'fcontent': open(os.path.join(__location__, 'outstanding_remainder.xlsx'),'rb').read()
					}]

				message = """
						Dear Sir/Madam,<br>
						<br>
						&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Our records indicate that some payments on your account are still due. Please find details below. <br>
						<b>If the amount has already been paid, Please Share Transaction Reference No. with other details.</b> Otherwise, Please forward us the total amount <br>
						stated below.<br>
						<br>
						If you have any query regarding your account, Please contact us.
						<br>
						<br>
						"""
				
				# cust data passing to html template(email_template)
				cust = {'customer_outstanding':customer_outstanding_data,
						'customer_name': i['customer_name'],
						'sum_inv_amt': sum_inv_amt,
						'sum_paid_amt': sum_paid_amt,
						'sum_cn_amt': sum_cn_amt,
						'sum_outstanding_amt': sum_outstanding_amt
						}

				rtemplate = Environment(loader=FileSystemLoader(__location__))
				html_file = rtemplate.get_template('email_template.html')
				email_template = html_file.render(**cust)

				frappe.sendmail(
								recipients = i['email_id'],
								subject = "Outstanding Statement",
								sender = "accounts@dhuparbrothers.com",
								message = message + email_template,
								# message = message,
								attachments = attachments
							)
				# print('mail_sended to --> ',i)
				# logging.info(i)
				with open('outstanding_mail_log1.log','a') as f:
					f.write(str(datetime.today()))
					f.write(' - ')
					f.write(str(i))
					f.write('\n')

				frappe.db.commit()

@frappe.whitelist()
def outstanding_mail_scheduler():
	data = frappe.db.sql(f"""
							SELECT orm.name, ormi.customer_name,ormi.email_id FROM `tabOutstanding Remainder Mail` orm
							JOIN `tabOutstanding Remainder Mail Items` ormi
							ON orm.name = ormi.parent
							WHERE
							orm.docstatus = 1
							AND
							orm.enable = 1
						""",as_dict=1)
	print('Outstanding mail: ',data)

	for i in data:
		orm_doc = frappe.get_doc('Outstanding Remainder Mail',i.name)
		send_outstanding_mail(orm_doc,event="weekly")

# CALL --> remianders/public/js/email_update.js
@frappe.whitelist()
def update_email_id(customer_name1,billing_email_id1=''):
	frappe.db.set_value('Customer',customer_name1,'billing_email_id',billing_email_id1)


# CALL --> remianders/public/js/dispatch_order.js
@frappe.whitelist()
def fetch_dispatch_data(name):
	so_data = frappe.db.sql(f"""
								SELECT modified as date, name, po_no, customer
								FROM
								 `tabSales Order`
								WHERE
								name = '{name}'
							""",as_dict=1)
	data = {}
	if len(so_data)!=0:
		for i in so_data:
			data['name'] = i.name
			data['customer'] = i.customer

			if i.po_no == None:
				data['po_no'] = 'Not Available'
			else:
				data['po_no'] = i.po_no

			data['date']=i.date

	# print(data)
	return data

# This function Fetch The mail from `tabAddress` using address_title and filter the mail and send outstanding mail
# n8n (Workflow Name - Outstandig Mail (EveryMonth))
@frappe.whitelist()
def filter_mail_and_send_outstanding_mail(**kwargs):
	# customer = frappe.db.get_list('Customer',
	# 		       filters={
	# 					'disabled': False	
	# 				})
	# print('customer: ',customer)

	customer = list(kwargs.values())

	# with open('outstanding_mail_log1.log','a') as f:
	# 	f.write(str('Batch'))
	# 	f.write(str(customer))
	# 	f.write('\n')

	for i in customer:
		get_email = frappe.db.sql(f"""
									select address_title as customer,email_id from `tabAddress`
									where address_title = "{i}"
								""",as_dict=1)
	
		email = set()
		for i in get_email:
			email.add(i.email_id)

		for unique_email in email:
			if unique_email != None:
				new_mail = unique_email.replace(' ','') # removed space from email
				new_mail2 = new_mail.replace(';',',') # removed semicolan from email

				if len(new_mail2) != 0:
					send_outstanding_mail(i.customer,new_mail2)
					# send_outstanding_mail(i.customer,"ajaypato.com@gmail.com")


# This code is for to send mail manually by calling this function and passing cutomer
@frappe.whitelist()
def filter_mail_and_send_outstanding_mail():

	# customer = list(kwargs.values())

	customer = ['SHREE HARI CONTROLS']

	for i in customer:
		get_email = frappe.db.sql(f"""
									select address_title as customer,email_id from `tabAddress`
									where address_title = "{i}"
								""",as_dict=1)

		email = set()
		for i in get_email:
			email.add(i.email_id)

		for unique_email in email:
			if unique_email != None:
				new_mail = unique_email.replace(' ','') # removed space from email
				new_mail2 = new_mail.replace(';',',') # removed semicolan from email

				if len(new_mail2) != 0:
					send_outstanding_mail(i.customer,new_mail2)
					# send_outstanding_mail(i.customer,"ajaypato.com@gmail.com")
 